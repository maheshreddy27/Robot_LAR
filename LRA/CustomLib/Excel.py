import xlrd
import xlwt
import openpyxl as opx
from openpyxl import Workbook
from openpyxl import load_workbook

from xlutils.copy import copy

def read_excelRow_file(path,sheetIndex,rowNumber):
    """
    Open and read an Excel file
    """
    book = xlrd.open_workbook(path)
    # print number of sheets
    print book.nsheets
    # print sheet names
    print book.sheet_names()
    # get the first worksheet
    first_sheet = book.sheet_by_index(int(sheetIndex))
    # read a row
    return  first_sheet.row_values(int(rowNumber))
    
def Get_Row_Count(path,sheetIndex):
	book = xlrd.open_workbook(path)
	print book.sheet_names()
	first_sheet = book.sheet_by_index(int(sheetIndex))
	print  first_sheet.nrows
	return  first_sheet.nrows
	
def write_value_to_File(path,row1,col,val):
 	wb = opx.load_workbook(path)
 	sheet = wb.active
 	print sheet
 	c1 = sheet.cell(row = 2, column = 1)
 	c1.value = val
 	wb.save(path)
 	print val, ' Datasheet is updated'



